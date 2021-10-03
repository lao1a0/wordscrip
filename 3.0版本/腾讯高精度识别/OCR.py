

import os
import re
import json
import time
import base64 
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from tencentcloud.common import credential
from tencentcloud.common.profile.client_profile import ClientProfile
from tencentcloud.common.profile.http_profile import HttpProfile
from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
from tencentcloud.ocr.v20181119 import ocr_client, models

def getFileList(dir,Filelist, ext=None):
    newDir = dir
    if os.path.isfile(dir):
        if ext is None:
            Filelist.append(dir)
        else:
            if ext in dir[-3:]:
                Filelist.append(dir)
    
    elif os.path.isdir(dir):
        for s in os.listdir(dir):
            newDir=os.path.join(dir,s)
            getFileList(newDir, Filelist, ext)
    return Filelist

def OCR():
    for imgpath in imglist:
        imgname= os.path.splitext(os.path.basename(imgpath))[0]
        xlsx1(imgname)
        #对图片进行base64编码
        with open(imgpath, 'rb') as f:
            base64_data = base64.b64encode(f.read())
            s = base64_data.decode()
            image = "data:image/jpeg;base64," + s
        #调用腾讯云api服务进行ocr识别
        try:
            cred = credential.Credential("AKIDNXSELDoQQCg4ADptZx5wMWcldFTjfqni", "oLcc5GAyf46mESZUL7ankQpxoVP8ML6Q")
            httpProfile = HttpProfile()
            httpProfile.endpoint = "ocr.tencentcloudapi.com"

            clientProfile = ClientProfile()
            clientProfile.httpProfile = httpProfile
            client = ocr_client.OcrClient(cred, "ap-beijing", clientProfile)

            req = models.GeneralAccurateOCRRequest()
            params = {
                "ImageBase64": image
            }
            req.from_json_string(json.dumps(params))
            resp = client.GeneralAccurateOCR(req)
            #print(resp.to_json_string())
            try:
                for i in range(0,300):
                    if json.loads(resp.to_json_string())['TextDetections'][i]['Confidence'] != None:
                        string = json.loads(resp.to_json_string())['TextDetections'][i]['DetectedText']
                        s = re.sub(u"([^\u4e00-\u9fa5\u0030-\u0039\u0041-\u005a\u0061-\u007a])","",string)   #正则除掉括号及其他字符
                        text = re.sub('[a-zA-Z0-9]', "", s)                                                #正则过滤掉数字和字母
                        print(text)
                        j = i + 1
                        xlsx2(imgname,j,text)
            except Exception as e:
                pass

        except TencentCloudSDKException as err:
            print(err)

def xlsx1(page):
    #读取数据
    wb2=openpyxl.load_workbook('1.xlsx')
    #创建sheet页
    wb2.create_sheet(page)
    #保存数据
    wb2.save('1.xlsx')
    #关闭excel
    wb2.close()

def xlsx2(page,i,data):
    #读取数据
    wb2=openpyxl.load_workbook('1.xlsx')
    #获取sheet页
    sheet2=wb2.get_sheet_by_name(page)
    #写入excel
    sheet2.cell(i,1).value = data
    #保存数据
    wb2.save('1.xlsx')
    #关闭excel
    wb2.close()

if __name__ == '__main__':
    #更改为存有图片的根目录
    org_img_folder = "C:/Users/dell/Desktop/腾讯高精度识别/image"
    imglist = getFileList(org_img_folder, [], 'png')
    print('一共发现了 '+str(len(imglist))+' 张图像\n')
    print('开始进行OCR识别并存入excel中\n')
    OCR()
    print("全部图片已处理完毕")
    time.sleep(120)

